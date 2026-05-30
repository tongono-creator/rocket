# -*- coding: utf-8 -*-
"""review.py — generate รูปรีวิวสินค้าจาก review_products.xlsx แล้วโพส FB"""

import sys, io, os, base64, requests, time, random, re
from datetime import datetime, timezone, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from google import genai
import openpyxl
from overlay_utils import add_overlay

GOOGLE_API_KEY    = os.environ.get("GOOGLE_API_KEY",    "")
PAGE_ACCESS_TOKEN = os.environ.get("PAGE_ACCESS_TOKEN", "")
PAGE_ID           = os.environ.get("PAGE_ID", "111830598532037")
TEXT_MODELS       = ["gemini-2.5-flash", "gemini-3.5-flash"]
OUTPUT_DIR        = "output"
EXCEL_PATH        = os.path.join(os.path.dirname(__file__), "review_products.xlsx")
ACCENT_COLOR      = (255, 215, 0) # Gold สำหรับ Rocket21

if not GOOGLE_API_KEY:
    try:
        from config import GOOGLE_API_KEY, PAGE_ACCESS_TOKEN
    except ImportError:
        pass

os.makedirs(OUTPUT_DIR, exist_ok=True)
client = genai.Client(api_key=GOOGLE_API_KEY, http_options={'timeout': 90.0})

def contains_thai(text):
    if not text:
        return False
    return bool(re.search(r'[\u0e00-\u0e7f]', text))

def segment_thai_text(text, client=client):
    if not text or not contains_thai(text):
        return text
    prompt = (
        "You are an expert Thai word segmentation tool. "
        "Your task is to insert a zero-width space character (\\u200b) at every natural word boundary in the provided Thai text. "
        "Strict rules:\n"
        "1. Do NOT modify, delete, or add any words, characters, punctuation, spaces, or newlines of the original text. "
        "Keep the exact same characters and layout.\n"
        "2. Do NOT add any introductory or concluding remarks. Output ONLY the segmented text.\n"
        "3. Ensure words like 'หวยออก', 'เงินเก็บ', 'แสนแรก', 'ทำงาน' are segmented at their natural boundaries (e.g., 'หวย\\u200bออก' or left as 'หวยออก', but never break syllables awkwardly).\n\n"
        f"Text to segment:\n{text}"
    )
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            segmented = resp.text.strip().replace('\\u200b', '\u200b')
            clean_orig = text.replace('\u200b', '').replace('\\u200b', '')
            clean_seg = segmented.replace('\u200b', '').replace('\\u200b', '')
            if len(clean_orig) == len(clean_seg):
                return segmented
        except Exception as e:
            print(f"[{model}] segment_thai_text failed: {e}")
    return text

def load_next_product():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=False):
        no    = row[0].value
        detail= row[1].value
        shopee= row[2].value
        lazada= row[3].value
        imgurl= row[4].value
        promo = row[5].value
        posted= row[6].value

        # ข้าม sample row และ posted แล้ว
        if not detail or "วางรายละเอียด" in str(detail):
            continue
        if str(posted).strip().lower() == "done" or str(posted).strip().startswith("done"):
            continue
        if not shopee or "xxx" in str(shopee):
            continue

        return {
            "row": row[0].row,
            "no": no,
            "detail": str(detail).strip(),
            "shopee": str(shopee).strip(),
            "lazada": str(lazada).strip() if lazada else "",
            "image_url": str(imgurl).strip() if imgurl else "",
            "promo": str(promo).strip() if promo else "",
        }, wb, ws
    return None, wb, ws

def mark_posted(wb, ws, row_num):
    bkk = timezone(timedelta(hours=7))
    ts = datetime.now(bkk).strftime("%Y-%m-%d %H:%M")
    ws.cell(row=row_num, column=7, value=f"done {ts}")
    wb.save(EXCEL_PATH)
    print(f"Marked row {row_num} as done")

def clean_promo(raw):
    """เอาเฉพาะบรรทัดที่มี ฿ หรือ ลด หรือ % หรือ ส่งฟรี"""
    if not raw:
        return ""
    lines = raw.strip().splitlines()
    kept = [l.strip() for l in lines if re.search(r'฿|ลด|%|ส่งฟรี|flash|sale', l, re.IGNORECASE)]
    return " | ".join(kept[:3]) if kept else ""

def extract_highlights(detail, promo):
    """ให้ AI สกัดจุดเด่นจาก raw detail"""
    prompt = (
        f"จากรายละเอียดสินค้านี้:\n{detail}\n\n"
        f"สกัดออกมาเป็น bullet points ภาษาไทยสั้นๆ 3-5 จุดเด่น "
        f"เน้นประโยชน์ที่คนซื้อสนใจ ห้ามใส่ข้อมูลราคาหรือโปรโมชั่น "
        f"ตอบแค่ bullet points เท่านั้น"
    )
    highlights = None
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            highlights = resp.text.strip()
            if highlights:
                break
        except Exception as e:
            print(f"[{model}] highlights failed: {str(e)[:80]}")
            
    if not highlights:
        print("[Warning] Highlights generation failed on all models. Falling back to local heuristic extraction.")
        lines = [l.strip() for l in detail.splitlines() if l.strip()]
        thai_lines = [l for l in lines if contains_thai(l)]
        if not thai_lines:
            thai_lines = lines
        
        bullets = []
        for line in thai_lines:
            cleaned = re.sub(r'^[•\-\*\d\.\s\u2013]+', '', line).strip()
            if cleaned and len(cleaned) > 5 and len(cleaned) < 100:
                bullets.append(f"• {cleaned}")
            if len(bullets) >= 4:
                break
        if not bullets:
            bullets = [f"• {line[:80]}" for line in thai_lines[:3]]
        highlights = "\n".join(bullets) if bullets else "• รายละเอียดเพิ่มเติมตามระบุในลิงก์ร้านค้าครับ"

    if promo:
        highlights += f"\n🔥 โปรตอนนี้: {promo}"
    return highlights

def download_image(url):
    """Download รูปแรกจาก URL"""
    first_url = url.strip().split("\n")[0].strip()
    resp = requests.get(first_url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    if resp.status_code == 200:
        bkk = timezone(timedelta(hours=7))
        ts = datetime.now(bkk).strftime("%Y%m%d_%H%M%S")
        ext = "webp" if "webp" in first_url else "jpg"
        path = os.path.join(OUTPUT_DIR, f"product_{ts}.{ext}")
        with open(path, "wb") as f:
            f.write(resp.content)
        print(f"Product image downloaded: {path}")
        return path
    raise RuntimeError(f"Image download failed: {resp.status_code}")

def generate_hook(detail, highlights):
    """สร้างหัวข้อสั้นพาดหัวรูปภาพ 2 บรรทัด คั่นด้วย '|'"""
    prompt = (
        f"จากรายละเอียดสินค้าต่อไปนี้:\n{detail}\n\n"
        f"จุดเด่นสินค้าที่สกัดแล้ว:\n{highlights}\n\n"
        "กรุณาสร้างคำพาดหัวโฆษณารีวิวสินค้านี้ภาษาไทยสั้นๆ 2 บรรทัด คั่นด้วยเครื่องหมาย pipe '|' (บรรทัด 1 | บรรทัด 2)\n"
        "กฎในการร่าง:\n"
        "- บรรทัด 1: คำโปรย/ชื่อเล่นสุดปังสไตล์วัยรุ่นหรือคนทำงานขำขัน (3-5 คำ)\n"
        "- บรรทัด 2: เหตุผลโดนใจ/จุดเด่นในการแก้ปัญหา (4-7 คำ)\n"
        "- ห้ามใช้เครื่องหมายคำพูด อัญประกาศ หรือข้อความนำหน้า/ตามหลังใดๆ\n"
        "- ห้ามมี Emoji ปนในหัวข้อนี้เด็ดขาด\n"
        "ตัวอย่าง: เบาะรองนั่งสู้ชีวิต | นั่งทำงาน 10 ชม. ไม่ปวดหลัง"
    )
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            result = resp.text.strip()
            label_pattern = r'^(ข้อความในโพสต์\s*Facebook|Facebook\s*Caption|Facebook\s*caption|Caption|caption|ข้อความบนรูป|ข้อความในรูป|ข้อความ|คำบรรยาย|คำอธิบาย|บรรทัดที่\s*\d+|บรรทัด\s*\d+|ประโยคที่\s*\d+|ประโยค\s*\d+|Hook\s*text|Hook|Line\s*\d+|[L|l]ine\s*\d+|\d+)\s*[:\-\.\s]\s*'
            result = re.sub(label_pattern, '', result, flags=re.IGNORECASE).strip()
            result = result.strip('"\'“”‘’')
            if "|" in result:
                parts = result.split("|", 1)
                line1 = parts[0].strip()
                line2 = parts[1].strip()
                return line1, line2
            else:
                return result[:15], ""
        except Exception as e:
            print(f"[{model}] hook generation failed: {e}")
    return "สินค้าแนะนำ", ""

def generate_caption(detail, shopee, lazada, promo, highlights):
    promo_line = f"\n🔥 โปรโมชั่น: {promo}" if promo else ""
    lazada_line = f"\n🛍️ Lazada → {lazada}" if lazada and "xxx" not in lazada else ""
    prompt = (
        f"เขียน Facebook post ภาษาไทยรีวิวสินค้าอย่างตรงไปตรงมาและน่าอ่าน สไตล์เพจรีวิวชื่อดัง (เช่น CatDumb)\n"
        f"เขียนด้วยบุคลิกแอดมินผู้ชาย (ใช้คำลงท้ายว่า 'ครับ' และแทนตัวว่า 'ผม' หรือ 'พี่' เท่านั้น)\n"
        f"รายละเอียดสินค้า:\n{detail}\n\n"
        f"จุดเด่นสินค้า:\n{highlights}\n\n"
        f"คุณต้องเขียนรีวิวโดยใช้เทคนิค 3 ขั้นตอนดังนี้:\n"
        f"1. เปิดให้น่าสนใจ (Hook): ประโยคเปิดหัวพาดหัวเรื่องให้น่าตื่นเต้น จี้ใจ ดึงดูดให้อยากอ่านต่อทันที\n"
        f"2. เล่าให้เห็นภาพ (Vivid Storytelling): รีวิวคุณสมบัติเด่น ประสิทธิภาพการใช้งานจริง หรือผลลัพธ์หลังจากใช้แล้วให้เห็นภาพชัดเจน\n"
        f"3. ปิดจบต้องบอกว่า 'ควรทำอะไร' (Call to Action): บอกให้สั่งซื้อโดยการกดที่ตะกร้า ชี้เป้าลิงก์ หรือบอกว่ามีโปรโมชั่นจัดหนักอยู่\n\n"
        f"เขียนให้น่าอ่าน สั้นกระชับ เป็นกันเอง ท้ายโพสต์ใส่แฮชแท็ก 2-3 อัน\n"
        f"ห้ามใช้ markdown ตัวหนา (**) และตอบเฉพาะตัวแคปชั่นรีวิวเท่านั้น"
    )
    caption = None
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            caption = resp.text.strip()
            if caption:
                break
        except Exception as e:
            print(f"[{model}] caption generation failed: {e}")
            
    if not caption:
        print("[Warning] Caption generation failed on all models. Falling back to local heuristic caption.")
        first_few_lines = " ".join([l.strip() for l in detail.splitlines() if l.strip()][:3])
        caption = (
            f"สวัสดีครับทุกคน วันนี้ผมมีสินค้าดีๆ มาแนะนำครับ!\n\n"
            f"ตัวนี้คือ {first_few_lines} บอกเลยว่าตอบโจทย์ชีวิตประจำวันมากครับ ลองดูรายละเอียดและจุดเด่นด้านล่างได้เลยครับ\n\n"
            f"{highlights}\n\n"
            f"ใครที่กำลังมองหาอยู่หรืออยากสั่งซื้อไปลองใช้งาน สามารถกดสั่งได้ที่ลิงก์ตะกร้าด้านล่างนี้ได้เลยครับผม 👇"
        )
    else:
        lines = caption.splitlines()
        while lines and (
            re.search(r'^(ได้เลย|นี่คือ|แน่นอน|โพสต์รีวิว|ครับ|ค่ะ|---)', lines[0].strip(), re.IGNORECASE)
            or lines[0].strip() in ("", "---")
        ):
            lines.pop(0)
        caption = "\n".join(lines).strip()
        
    caption += f"{promo_line}\n\n👉 Shopee → {shopee}{lazada_line}"
    return caption

def post_to_page(img_path, caption):
    print("Posting to Facebook Page...")
    with open(img_path, "rb") as f:
        resp = requests.post(
            f"https://graph.facebook.com/v25.0/{PAGE_ID}/photos",
            data={"access_token": PAGE_ACCESS_TOKEN, "caption": caption, "published": "true"},
            files={"source": ("review.png", f, "image/png")},
            timeout=60
        )
    result = resp.json()
    if "id" in result:
        post_id = result.get("post_id") or result["id"]
        print(f"Page Posted! ID: {post_id}")
        print(f"https://www.facebook.com/{post_id}")
        return post_id
    else:
        print(f"FB Error: {result}")
        raise SystemExit(1)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Run without posting or marking as done")
    args = parser.parse_args()

    product, wb, ws = load_next_product()
    if not product:
        print("ไม่มีสินค้าที่ต้องโพส (ครบแล้วหรือยังไม่ได้เพิ่ม)")
        raise SystemExit(0)

    print(f"Product #{product['no']}: {product['detail'][:60]}...")

    promo_clean = clean_promo(product["promo"])
    highlights  = extract_highlights(product["detail"], promo_clean)
    print(f"Highlights:\n{highlights}\n")

    line1, line2 = generate_hook(product["detail"], highlights)
    line1 = segment_thai_text(line1, client)
    line2 = segment_thai_text(line2, client)
    print(f"Hook generated: {line1} | {line2}")

    product_img = download_image(product["image_url"])
    
    # PIL Overlay
    try:
        review_img = add_overlay(product_img, line1, line2, ACCENT_COLOR)
        os.unlink(product_img)
        print(f"Review image overlaid: {review_img}")
    except Exception as overlay_err:
        print(f"Overlay failed, using original image: {overlay_err}")
        review_img = product_img

    caption = generate_caption(
        product["detail"], product["shopee"],
        product["lazada"], promo_clean, highlights
    )
    print(f"Caption:\n{caption}\n")

    if args.dry_run:
        print(f"Dry run complete. Local image path: {review_img}")
    else:
        post_to_page(review_img, caption)
        if os.path.exists(review_img):
            os.unlink(review_img)
        mark_posted(wb, ws, product["row"])
